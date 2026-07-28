"""Excel 月报生成。"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.services import stats_service
from app.services.lesson_service import list_lessons

HEADER_FILL = PatternFill(start_color="4C7DFF", end_color="4C7DFF", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def _set_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER


def generate_xlsx(db: Session, from_date: date, to_date: date) -> bytes:
    wb = Workbook()

    # Sheet 1: 课时明细
    ws1 = wb.active
    ws1.title = "课时明细"
    _set_header(
        ws1,
        ["日期", "星期", "开始时间", "时长(小时)", "学生", "单价", "金额", "状态", "备注"],
    )
    week_names = ["一", "二", "三", "四", "五", "六", "日"]
    lessons = list_lessons(db, from_date=from_date, to_date=to_date)
    for ls in lessons:
        ws1.append(
            [
                ls.date.isoformat(),
                "周" + week_names[ls.date.weekday()],
                ls.start_time,
                ls.duration_hours,
                ls.student.name if ls.student else "",
                ls.student.hourly_rate if ls.student else "",
                ls.price,
                ls.status,
                ls.note or "",
            ]
        )
    for col_letter in "ABCDEFGHI":
        ws1.column_dimensions[col_letter].width = 14

    # Sheet 2: 学生汇总
    ws2 = wb.create_sheet("学生汇总")
    _set_header(
        ws2,
        ["学生", "节数", "总课时", "总收入", "请假次数", "调课次数"],
    )
    ranking = stats_service.student_ranking(db, from_date=from_date, to_date=to_date)
    for r in ranking:
        ws2.append(
            [
                r.name,
                r.lesson_count,
                r.total_hours,
                r.total_income,
                r.leave_count,
                r.reschedule_count,
            ]
        )
    for col_letter in "ABCDEF":
        ws2.column_dimensions[col_letter].width = 14

    # Sheet 3: 日收入汇总
    ws3 = wb.create_sheet("日收入汇总")
    _set_header(ws3, ["日期", "收入", "课时", "节数"])
    rng = stats_service.range_stats(
        db, from_date=from_date, to_date=to_date, granularity="day"
    )
    for b in rng.buckets:
        ws3.append([b.bucket, b.income, b.hours, b.lesson_count])
    ws3.append([])
    ws3.append(["合计", rng.total_income, rng.total_hours, rng.total_lessons])
    for col_letter in "ABCD":
        ws3.column_dimensions[col_letter].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
